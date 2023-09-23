import openpyxl

def copiar_libro(origen, destino, criterios):
  """
  Copia un libro de origen en un libro de destino, con todas sus celdas.

  Args:
    origen: Ruta del archivo de Excel de origen.
    destino: Ruta del archivo de Excel de destino.
    criterios: Conjunto de criterios para copiar las celdas.

  Returns:
    Ninguno.
  """

  # Cargamos los libros de origen y destino.
  libro_origen = openpyxl.load_workbook(origen)
  libro_destino = openpyxl.Workbook()

  # Recorremos todas las hojas del libro de origen.
  for hoja_origen in libro_origen.worksheets:
    # Añadimos una nueva hoja al libro de destino con el mismo nombre.
    hoja_destino = libro_destino.create_sheet(hoja_origen.title)

    # Recorremos todas las celdas de la hoja de origen.
    for fila in hoja_origen.iter_rows():
      # Copiamos las celdas que cumplan los criterios.
      for celda in fila:
        if celda.value in criterios:
          hoja_destino[celda.row][celda.column] = celda.value

  # Guardamos el libro de destino.
  libro_destino.save(destino)


if __name__ == "__main__":
  # Rutas de los archivos de Excel.
  origen = "archivo_origen.xlsx"
  destino = "archivo_destino.xlsx"

  # Conjunto de criterios para copiar las celdas.
  criterios = {"1", "2", "3"}

  # Copiamos el libro de origen en el libro de destino, con las celdas que cumplan los criterios.
  copiar_libro(origen, destino, criterios)
