# Programa para leer y reemplazar el contenido de una hoja de Excel

import openpyxl


def leer_hoja_origen(archivo_origen, hoja_origen):
  """
  Lee el contenido de la hoja de Excel especificada en el archivo de origen.

  Args:
    archivo_origen: El archivo de origen.
    hoja_origen: El nombre de la hoja de Excel a leer.

  Returns:
    Una lista con las filas de la hoja de Excel.
  """

  # Abrimos el archivo de origen
  libro_origen = openpyxl.load_workbook(archivo_origen)

  # Obtenemos la hoja de Excel especificada
  hoja_origen = libro_origen.get_sheet_by_name(hoja_origen)

  # Leemos el contenido de la hoja de Excel
  filas = []
  for fila in hoja_origen.iter_rows():
    filas.append([celda.value for celda in fila])

  # Cerramos el archivo de origen
  libro_origen.close()

  return filas


def reemplazar_hoja_destino(archivo_destino, hoja_destino, filas):
  """
  Reemplaza el contenido de la hoja de Excel especificada en el archivo de destino.

  Args:
    archivo_destino: El archivo de destino.
    hoja_destino: El nombre de la hoja de Excel a reemplazar.
    filas: Las filas con el nuevo contenido de la hoja de Excel.
  """

  # Abrimos el archivo de destino
  libro_destino = openpyxl.load_workbook(archivo_destino)

  # Obtenemos la hoja de Excel especificada
  hoja_destino = libro_destino.get_sheet_by_name(hoja_destino)

  # Reemplazamos el contenido de la hoja de Excel
  for fila_index, fila in enumerate(filas):
    for celda_index, celda in enumerate(fila):
      hoja_destino.cell(fila_index + 1, celda_index + 1).value = celda

  # Guardamos el archivo de destino
  libro_destino.save(archivo_destino)


def main():
  # Solicitamos los nombres de los archivos de origen y destino
  archivo_origen = input("Introduce el nombre del archivo de origen: ")
  hoja_origen = input("Introduce el nombre de la hoja de origen: ")
  archivo_destino = input("Introduce el nombre del archivo de destino: ")
  hoja_destino = input("Introduce el nombre de la hoja de destino: ")

  # Leemos el contenido de la hoja de Excel de origen
  filas = leer_hoja_origen(archivo_origen, hoja_origen)

  # Reemplazamos el contenido de la hoja de Excel de destino
  reemplazar_hoja_destino(archivo_destino, hoja_destino, filas)


if __name__ == "__main__":
  main()
